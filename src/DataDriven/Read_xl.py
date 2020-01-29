'''
Created on 21-Jan-2020

@author: praveen
'''

import openpyxl as opp

path = "/Users/praveen/Documents/Eclipse Project/PythonAutomation/SeleniumPython/DATA/DDT.xlsx"

workbook = opp.load_workbook(path)

sheet = workbook.active  

#sheet = workbook.get_sheet_by_name("driven")

rows = sheet.max_row

columns = sheet.max_column

print(rows)

print(columns)
 
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(row=r,column=c).value,end = "   ")
        
    print()
        
    